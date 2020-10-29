import base64
import logging
import os
import re
import string
import tempfile
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, AccessError, UserError
from openpyxl.styles import PatternFill, Border, Alignment, Font


_logger = logging.getLogger(__name__)
try:
    import openpyxl
except ImportError:
    msg = _('Install python module "openpyxl" in order to create Excel documents')
    raise ValidationError(msg)
try:
    import csv
except ImportError:
    msg = _('Install python module "csv" in order to generate CSV')
    raise ValidationError(msg)


class PayrollReports(models.Model):
    _inherit = 'hr.payslip.run'

    def GetNetPay(self):
        super().GetNetPay()

        employer_bank_account = self.env['res.config.settings'].search(
            []).mapped('employer_bank_account')

        for rec in self:
            if rec.slip_ids:
                filename_netpay = 'NET_PAY-' + re.sub(
                    '[^A-Za-z0-9]+', '', rec.name) + '_' + fields.Datetime.context_timestamp(
                    self, fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.xlsx'
                wb = openpyxl.Workbook()
                ws = wb.active
                fr = 2  # First row of data

                cols = [
                    'BENEFICIARY NAME',
                    'BENEFICIARY A/C NUMBER',
                    'PAYMENT AMOUNT',
                    'BANK NAME',
                    'BANK CODE,BRANCH CODE',
                    'PAYMENT DETAILS',
                    'YOUR COMPANY DEBIT ACCOUNT NUMBER',
                    'PURPPOSE OF PAYMENT CODE']
                # DATA HEADERS
                # 'ABCDEFGH'
                for k, x in enumerate(string.ascii_uppercase[0:8]):
                    ws[x + str(fr - 1)] = cols[k]

                for key, slip in enumerate(rec.slip_ids):
                    ws['A' + str(fr + key)] = slip.employee_id.name or None
                    ws['B' + str(fr + key)
                       ] = slip.employee_id.employee_bank_account or None
                    ws['C' + str(fr + key)] = slip.line_ids.search([(
                        'salary_rule_id', '=', rec.env.ref(
                            'hr_ke.ke_rule120').id),
                        ('slip_id', '=', slip.id)],
                        limit=1).total  # Total Net Pay

                    ws['D' + str(fr + key)
                       ] = slip.employee_id.bank_name or None
                    ws['E' + str(fr + key)
                       ] = slip.employee_id.branch_code or None
                    ws['F' + str(fr + key)] = slip.payslip_run_id.name or None
                    ws['G' + str(fr + key)] = employer_bank_account[0] if len(
                        employer_bank_account) > 0 else None
                    ws['H' + str(fr + key)] = '048'

                xls_path = self.env['hr.ke'].create_xls()
                wb.save(xls_path)
                rec.env['hr.ke'].save_attachment(
                    filename_netpay, xls_path, self._name, rec.id)
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)

    def GetPayrollSummary(self):

        wb = openpyxl.Workbook()
        ws = wb.active

        col = [
            'SERIAL',
            'PAYROLL NO',
            'LAST NAME',
            'FIRST NAME',
            'DEPARTMENT',
            'BASIC PAY',
            'HOUSE ALLOWANCES',
            'LEAVE',
            'BONUS',
            'OVERTIME ALLOWANCES',
            'OTHER ALLOWANCES',
            'ABSENT AMOUNT',  # -----
            'GROSS PAY',
            'BENEFITS',
            'TAXABLE PAY',
            'P.A.Y.E',
            'N.S.S.F',
            'N.H.I.F',
            'H.E.L.B',
            # 'ALLOWED DEDUCTIONS',
            'SALARY ADVANCE',
            'SACCO',
            'PLEDGE',
            'LOAN',
            # 'NET TAXABLE PAY',
            'TOTAL DEDUCTIONS',
            'NETPAY'
        ]
        data = []
        for cell, title in zip(list(string.ascii_uppercase), col):
            ws[f'{cell}6'] = title

        # EXTRA CELLS
        # ws['AA6'] = 'NET PAY'

        for rec in self:
            if rec.slip_ids:
                filename_summary = 'Payroll_Summary-' + re.sub(
                    '[^A-Za-z0-9]+', '', rec.name) + '_' + fields.Datetime.context_timestamp(
                    self, fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.xlsx'
                ws['A1'] = rec.journal_id.company_id.name or None
                ws['A2'] = 'PAYROLL SUMMARY'
                ws['B2'] = rec.name or None
                ws['A3'] = 'DATE FROM'
                ws['B3'] = rec.date_start or None
                ws['A4'] = 'DATE TO'
                ws['B4'] = rec.date_end or None

                for slip in rec.slip_ids:

                    helb = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule107').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0  # Total  HELB
                    nhif = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule106').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    nssf = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule55').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    total_net_paye = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule105').id), ('slip_id', '=', slip.id)], limit=1).total
                    total_net_taxable = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule85').id), ('slip_id', '=', slip.id)], limit=1).total
                    total_allowed_deds = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule80').id), ('slip_id',  '=', slip.id)], limit=1).total
                    total_gross_pay = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule30').id),  ('slip_id', '=', slip.id)],  limit=1).total or 0.0
                    total_taxable = slip.line_ids.search([('salary_rule_id', '=',  rec.env.ref(
                        'hr_ke.ke_rule45').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    house_allowance = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule17').id), ('slip_id', '=', slip.id)],  limit=1).total or 0.0
                    basic_pay = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule10').id), ('slip_id', '=', slip.id)], limit=1).total
                    leave = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule15').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    bonus = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule11').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    overtime = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule13').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    other_allowances = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule19').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    advance = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule108').id), ('slip_id', '=', slip.id)], limit=1).total
                    absent = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule111').id), ('slip_id', '=', slip.id)], limit=1).total
                    sacco = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule109').id), ('slip_id', '=', slip.id)], limit=1).total
                    pledge = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule110').id), ('slip_id', '=', slip.id)], limit=1).total
                    loan = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule112').id), ('slip_id', '=', slip.id)], limit=1).total
                    total_net_taxable = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule85').id), ('slip_id', '=', slip.id)], limit=1).total
                    total_net_paye = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule105').id), ('slip_id', '=', slip.id)], limit=1).total
                    total_deductions = advance + sacco + pledge + \
                        loan + total_net_paye + nssf + nhif + helb
                    total_net_pay = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule120').id), ('slip_id', '=', slip.id)], limit=1).total

                    data.append({
                        'payroll_no': slip.employee_id.payroll_no,
                        'last_name': slip.employee_id.display_name.split(' ')[-1] or None,
                        'first_name': slip.employee_id.display_name.split(' ')[0] or None,
                        'department': slip.employee_id.department_id.name or None,

                        'basic_pay': basic_pay,
                        'house_all': house_allowance,
                        'leave': leave,
                        'bonus': bonus,
                        'overtime_all': overtime,
                        'other_allowances': other_allowances,
                        'gross_pay': total_gross_pay,
                        'taxable_pay': total_taxable,
                        'total_net_taxable': total_net_taxable,
                        'benefits': total_gross_pay-total_taxable,

                        'allowed_ded': total_allowed_deds,
                        'salary_adv': advance,
                        'abs_amount': absent,
                        'sacco': sacco,
                        'pledge': pledge,
                        'loan': loan,
                        'paye': total_net_paye,
                        'nssf': nssf,
                        'nhif': nhif,
                        'helb': helb,

                        'total_deductions': total_deductions,
                        'total_net_pay': total_net_pay,
                    })
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)

        data = sorted(data, key=lambda i: i['payroll_no'])

        seen = {}
        for key, val in enumerate(data):
            for i in val:
                if i != 'payroll_no' and i != 'last_name' and i != 'first_name' and i != 'department':
                    if i in seen:
                        seen[i] += data[key][i]
                    else:
                        seen[i] = data[key][i]

        for index, val in enumerate(data):
            row = index+7
            ws.cell(row, 1).value = index + 1
            ws.cell(
                row, 2).value = val['payroll_no'] if val['payroll_no'] != 0 else None
            ws.cell(row, 3).value = val['last_name']
            ws.cell(row, 4).value = val['first_name']
            ws.cell(row, 5).value = val['department']
            ws.cell(row, 6).value = val['basic_pay']
            ws.cell(row, 7).value = val['house_all']
            ws.cell(row, 8).value = val['leave']
            ws.cell(row, 9).value = val['bonus']
            ws.cell(row, 10).value = val['overtime_all']
            ws.cell(row, 11).value = val['other_allowances']
            ws.cell(row, 12).value = val['abs_amount']
            ws.cell(row, 13).value = val['gross_pay']
            ws.cell(row, 14).value = val['benefits']
            ws.cell(row, 15).value = val['taxable_pay']
            ws.cell(row, 16).value = val['paye']
            ws.cell(row, 17).value = val['nssf']
            ws.cell(row, 18).value = val['nhif']
            ws.cell(row, 19).value = val['helb']

            # ws.cell(row, 15).value = val['allowed_ded']
            ws.cell(row, 20).value = val['salary_adv']
            ws.cell(row, 21).value = val['sacco']
            ws.cell(row, 22).value = val['pledge']
            ws.cell(row, 23).value = val['loan']
            # ws.cell(row, 21).value = val['total_net_taxable']
            ws.cell(row, 24).value = val['total_deductions']
            ws.cell(row, 25).value = val['total_net_pay']

        ws[f'E{len(data)+8}'] = 'TOTAL'
        ws[f'F{len(data)+8}'] = seen['basic_pay']
        ws[f'G{len(data)+8}'] = seen['house_all']
        ws[f'H{len(data)+8}'] = seen['leave']
        ws[f'I{len(data)+8}'] = seen['bonus']
        ws[f'J{len(data)+8}'] = seen['overtime_all']
        ws[f'K{len(data)+8}'] = seen['other_allowances']

        ws[f'L{len(data)+8}'] = seen['abs_amount']

        ws[f'M{len(data)+8}'] = seen['gross_pay']
        ws[f'N{len(data)+8}'] = seen['benefits']
        ws[f'O{len(data)+8}'] = seen['taxable_pay']
        ws[f'P{len(data)+8}'] = seen['paye']
        ws[f'Q{len(data)+8}'] = seen['nssf']
        ws[f'R{len(data)+8}'] = seen['nhif']
        ws[f'S{len(data)+8}'] = seen['helb']

#         ws[f'O{len(data)+8}'] = seen['allowed_ded']
        ws[f'T{len(data)+8}'] = seen['salary_adv']
        ws[f'U{len(data)+8}'] = seen['sacco']
        ws[f'V{len(data)+8}'] = seen['pledge']
        ws[f'W{len(data)+8}'] = seen['loan']
#         ws[f'U{len(data)+8}'] = seen['total_net_taxable']

        ws[f'X{len(data)+8}'] = seen['total_deductions']
        ws[f'Y{len(data)+8}'] = seen['total_net_pay']
        for rec in self:
            if rec.slip_ids:
                # Save file as attachment
                xls_path = self.env['hr.ke'].create_xls()
                wb.save(xls_path)
                rec.env['hr.ke'].save_attachment(
                    filename_summary, xls_path, self._name, rec.id)
