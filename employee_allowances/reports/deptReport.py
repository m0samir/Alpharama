from odoo import models, _
import openpyxl as xl
from odoo.exceptions import ValidationError
import json
import string
from datetime import datetime
import re


class DepartmentalReport(models.Model):
    _inherit = 'hr.payslip.run'

    def deptReport(self):
        wb = xl.Workbook()
        ws = wb.active
        data = {}
        fr = 6
        # STYLES
        big_font = xl.styles.Font(size=15, bold=True)
        small_font = xl.styles.Font(size=13, bold=True)
        cols = ['PAYROLL NUMBER', 'NAME', 'BASIC', 'H/ALLOWANCE', 'OVERTIMR', 'ABSENT', 'GROSS PAY',
                'TAXABLE PAY', 'PAYE', 'NSSF', 'NHIF', 'ADVANCE', 'SACCO', 'LOAN', 'TOTAL DEDUCTIONS', 'NET PAY']

        for i, j in enumerate(string.ascii_uppercase[1:17]):
            ws[f'{j}{fr}'] = cols[i]

        for rec in self:
            if rec.slip_ids:
                month = rec.date_start.strftime('%m')
                year = rec.date_start.strftime('%Y')

                months = {'01': 'January', '02': 'February', '03': 'March', '04': 'April', '05': 'May', '06': 'June',
                          '07': 'July', '08': 'August', '09': 'September', '10': 'October', '11': 'November', '12': 'December'}

                ws['A1'] = rec.journal_id.company_id.name
                ws['A1'].font = big_font
                # ws['A1'].alignment = xl.styles.Alignment(
                #     horizontal='center', vertical='center')

                ws['A2'] = f'Company Payroll Report For The Month {months[month]}-{year}'
                ws['A2'].font = small_font

                ws['D4'] = 'EARNINGS'
                ws['D4'].font = big_font

                ws['J4'] = 'DEDUCTIONS'
                ws['J4'].font = big_font

                for slip in rec.slip_ids:
                    dept = self.env['hr.department'].search(
                        [('id', '=', slip.employee_id.department_id.id)]).mapped('name')

                    advance = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule108').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    helb = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule107').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0  # Total  HELB
                    nhif = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule106').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    nssf = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule55').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    sacco = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule109').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    loan = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule112').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    pledge = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule110').id), ('slip_id', '=', slip.id)], limit=1).total
                    paye = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule105').id), ('slip_id', '=', slip.id)], limit=1).total

                    vals = {
                        'pid': slip.employee_id.payroll_no,
                        'name': slip.employee_id.name,
                        'basic': slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                            'hr_ke.ke_rule10').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0,
                        'house': slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                            'hr_ke.ke_rule17').id), ('slip_id', '=', slip.id)],  limit=1).total or 0.0,
                        'overtime': slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                            'hr_ke.ke_rule13').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0,
                        'absent': slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                            'hr_ke.ke_rule111').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0,
                        'gross_pay': slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                            'hr_ke.ke_rule30').id),  ('slip_id', '=', slip.id)],  limit=1).total or 0.0,
                        'total_taxable': slip.line_ids.search([('salary_rule_id', '=',  rec.env.ref(
                            'hr_ke.ke_rule45').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0,
                        'paye': paye,
                        'nssf': nssf,
                        'nhif': nhif,
                        'advance': advance,
                        'sacco': sacco,
                        'loan': loan,
                        'total_deductions': advance + sacco + pledge + loan + paye + nssf + nhif + helb,
                        'net_pay': slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                            'hr_ke.ke_rule120').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0,
                    }
                    if dept[0] in data:
                        data[dept[0]].append(vals)
                    else:
                        data[dept[0]] = [vals]

                for i, key in enumerate(data):
                    fr += 1
                    ws.cell(fr, 1).value = key
                    ws.cell(fr, 1).font = small_font

                    for val in data[key]:
                        fr += 1
                        ws.cell(fr, 2).value = val['pid']
                        ws.cell(fr, 3).value = val['name']
                        ws.cell(fr, 4).value = val['basic']
                        ws.cell(fr, 5).value = val['house']
                        ws.cell(fr, 6).value = val['overtime']
                        ws.cell(fr, 7).value = val['absent']
                        ws.cell(fr, 8).value = val['gross_pay']
                        ws.cell(fr, 9).value = val['total_taxable']
                        ws.cell(fr, 10).value = val['paye']
                        ws.cell(fr, 11).value = val['nssf']
                        ws.cell(fr, 12).value = val['nhif']
                        ws.cell(fr, 13).value = val['advance']
                        ws.cell(fr, 14).value = val['sacco']
                        ws.cell(fr, 15).value = val['loan']
                        ws.cell(fr, 16).value = val['total_deductions']
                        ws.cell(fr, 17).value = val['net_pay']
                    fr += 1
                xls_path = self.env['hr.ke'].create_xls()
                wb.save(xls_path)
                rec.env['hr.ke'].save_attachment(
                    f'Department Report For {months[month]}-{year} on {datetime.now().strftime("%d_%m_%Y-%H_%M_%S")}', xls_path, self._name, rec.id)
            else:
                raise ValidationError(_("No Payslips to process!"))
